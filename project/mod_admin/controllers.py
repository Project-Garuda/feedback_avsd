from project.mod_faculty.models import Faculty,Courses,Feedback,Theory,Lab,Tutorial,UploadCourses,Admin
from project.mod_student.models import Section,Student,UploadSection
from project import db_session, bcrypt,app,Base,engine,create_engine_models,delete_engine_models
from flask import Flask, render_template, request, redirect, url_for, Blueprint, session,abort,flash
from werkzeug.utils import secure_filename
import os
import openpyxl as op

mod_admin = Blueprint('admin', __name__)

@mod_admin.route("/dashboard/", methods=['GET', 'POST'])#route for admin dashboard
def admin_dashboard():
    """This function is responsible for displaying content of the admin dashboard"""
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()#admin details are fetched
        upload_courses = UploadCourses.query.all()#All the active courses are fetched
        return render_template('admin/admin_dashboard.html',
        upload_courses=upload_courses,
        faculty = Faculty,
        course_names = Courses,
        admin = admin,
        status=app.config['feedback_status'],
        )#a html template is rendered with the data that is fetched
    else:
        return redirect(url_for('admin_home'))#if there is no admin we redirect to home

@mod_admin.route('/logout/')#route for admin logout
def logout():
    """If admin exists in the session admin is logged out and returned to home else just redirect to home"""
    if('admin' in session):
        session.pop('admin', None)
        flash('You have been logged out')
    return redirect(url_for('admin_home'))

@mod_admin.route("/view/<int:id>", methods=['GET', 'POST'])#route for viewing a particular course response
def view_responses(id):
    """This function is called when admin wishes to view responses of a particular course,this function will be
    called along with the id of upload courses for which which admin wants to view responses"""
    if 'admin' in session:#check if admin is in session or not
        my_obj = UploadCourses.query.filter(UploadCourses.id==id).first()#get object with that particular course
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if my_obj.course == 0:#if the course is theory course
            theory = Theory.query.filter(Theory.id == id).first()
            if theory is None:
                abort(404)
            theory_dict = theory.fetch_dict()
            theory_dict['no_responses'] = theory.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(Feedback.upload_courses_id==id).all()
            return render_template('admin/view_responses_theory.html',
            dict=theory_dict,
            admin=admin,
            remarks=remarks,)
        elif my_obj.course==1:#if the course is lab course
            lab = Lab.query.filter(Lab.id == id).first()
            if lab is None:
                abort(404)
            lab_dict = lab.fetch_dict()
            lab_dict['no_responses'] = lab.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(Feedback.upload_courses_id==id).all()#getting remarks
            return render_template('admin/view_responses_lab.html',
            dict=lab_dict,
            admin=admin,
            remarks=remarks,)
        else:
            #if the course is tutorial course
            tutorial = Tutorial.query.filter(Lab.id == id).first()
            if tutorial is None:
                abort(404)
            tutorial_dict = lab.fetch_dict()
            tutorial_dict['no_responses'] = tutorial.no_responses
            remarks = Feedback.query.with_entities(Feedback.remark).filter(Feedback.upload_courses_id==id).all()
            return render_template('admin/view_responses_tutorial.html',
            dict=tutorial_dict,
            admin=admin,
            remarks=remarks,)
    else:
        return redirect(url_for('admin_home'))

@mod_admin.route("/upload",methods=['GET','POST'])#uploading data route
def admin_upload():
    """"The function is responsible for taking the data from admin and updataing the database with new data
    collected from the admin."""
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method == 'POST':
            files = []
            try:
                for file in request.files:
                    current_file = request.files[file]
                    names=current_file.filename.split('.')
                    name=names[len(names)-1]
                    if current_file:
                        filename = secure_filename(file+'.'+name)
                        files.append(filename)
                        #upload folder is location where filed are stored locally
                        current_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                if process_data(files)==1:
                    #returning success message
                    flash('Data processed successfully')
                    return redirect(url_for('admin.admin_dashboard'))
                else:
                    #Error case
                    flash('Error')
                    return redirect(url_for('admin.admin_upload'))
            except Exception as e:
                print(e)
                flash('unknown error')
                return redirect(url_for('admin.admin_upload'))
        else:
            return render_template('admin/admin_upload.html',
            admin=admin,)
    else:
        return redirect(url_for('admin_home'))

def process_data(files):
    """Core function where all the data processing happens."""
    admin_objs = Admin.query.all()#To avoid removing admin credentials
    try:
        delete_engine_models()
        create_engine_models()
        for obj in admin_objs:
            db_session.add(Admin(obj.id,obj.name,obj.password))
        db_session.commit()
    except Exception as e:
        print(e)
        flash('Error In Clearing Data')
        return redirect(url_for('admin.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], files[0]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = int(sheet_obj.cell(row = i, column = 1).value)
        name = sheet_obj.cell(row = i, column = 2).value
        password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
        db_session.add(Faculty(id,name,password))
    try:
        db_session.commit()
    except Exception as e:
        flash('Error in processing Faculty data')
        return redirect(url_for('admin.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], files[1]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = int(sheet_obj.cell(row = i, column = 1).value)
        name = sheet_obj.cell(row = i, column = 2).value
        password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
        db_session.add(Student(id,name,password))
    try:
        db_session.commit()
    except Exception as e:
        flash('Error in processing Student data')
        return redirect(url_for('admin.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], files[2]))
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row+1):
        id = sheet_obj.cell(row = i, column = 1).value
        name = sheet_obj.cell(row = i, column = 2).value
        db_session.add(Courses(id,name))
    try:
        db_session.commit()
    except Exception as e:
        flash('Error in processing Course data')
        return redirect(url_for('admin.admin_upload'))

    wb_obj = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], files[3]))
    sheet_obj = wb_obj.active
    m_col = sheet_obj.max_column
    for i in range(1, m_col+1):
        id = sheet_obj.cell(row = 1, column = i).value
        db_session.add(Section(id))
        try:
            db_session.commit()
        except Exception as e:
            flash('Error in processing Upload sections data')
            return redirect(url_for('admin.admin_upload'))
        for j in range(2,m_row+1):
            obj = sheet_obj.cell(row = j, column = i).value
            if obj is None:
                break
            db_session.add(UploadSection(id,int(obj)))
    try:
        db_session.commit()
    except Exception as e:
        flash('Error in processing Upload sections data')
        return redirect(url_for('admin.admin_upload'))
    return 1

@mod_admin.route('/add',methods=['GET', 'POST'])
def add_user():
    """Function responsible for Add user feature."""
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method=='POST':
            id = int(request.form['id'])
            name = request.form['name']
            role = request.form['role']
            if role=='faculty':
                fac = Faculty.query.filter(Faculty.id == id).first()
                if fac is not None:
                    flash('Faculty already exists!')
                    return redirect(url_for('admin.add_user'))
                password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
                db_session.add(Faculty(id,name,password))
            else:
                section = request.form['section']
                stu = Student.query.filter(Student.id == id).first()
                if stu is not None:
                    flash('Student already exists!')
                    return redirect(url_for('admin.add_user'))
                if section is None:
                    flash('Please enter Section id!')
                    return redirect(url_for('admin.add_user'))
                sec = Section.query.filter(Section.id == section).first()
                if sec is None:
                    flash('Invalid Section ID')
                    return redirect(url_for('admin.add_user'))
                password = bcrypt.generate_password_hash(str(id)).decode('utf-8')
                db_session.add(Student(id,name,password))
                db_session.add(UploadSection(section,id))
            try:
                db_session.commit()
                flash('User added Successfully')
                return redirect(url_for('admin.admin_dashboard'))
            except Exception as e:
                flash('Unknown Error!')
                return redirect(url_for('admin.add_user'))
        else:
            return render_template('admin/add_user.html',
            admin=admin,)
    return redirect(url_for('admin_home'))

@mod_admin.route('/change',methods=['GET', 'POST'])
def change_password():
    """Function responsible for changing password of admin"""
    if 'admin' in session:
        admin = Admin.query.filter(Admin.id == session['admin']).first()
        if request.method == 'POST':
            if bcrypt.check_password_hash(admin.password, request.form['old_pass']):
                new_pass = bcrypt.generate_password_hash(request.form['new_pass']).decode('utf-8')
                update_admin = Admin.query.filter(Admin.id == session['admin']).update({'password':new_pass})
                try:
                    db_session.commit()
                except Exception as e:
                    print(e)
                    flash('Unknown error!')
                    return redirect(url_for('admin.change_password'))
                flash('Successfully updated the password')
                return redirect(url_for('admin.admin_dashboard'))
            else:
                flash('Old Password is incorrect')
                return redirect(url_for('admin.change_password'))
        else:
            return render_template('admin/change_password.html',
            admin=admin,)
    return redirect(url_for('admin_home'))

@mod_admin.route('/toggle',methods=['GET'])
def toggle_feedback():
    """Function responsible for Changing the status of feedback"""
    if 'admin' in session:
        app.config['feedback_status']=1-app.config['feedback_status']
        return redirect(url_for('admin.admin_dashboard'))
    return redirect(url_for('admin_home'))
